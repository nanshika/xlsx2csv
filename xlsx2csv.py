#!/usr/bin/python3
# -*- coding: UTF8 -*-
# python Export_all.py abc.xlsx

from optparse import OptionParser
import openpyxl
import re
import os
pwd = os.getcwd()

def export_all(file_path):
    ''' print all cells without duplication '''
    wb = openpyxl.load_workbook(file_path)
    file_name = re.sub(r'.+\\(.+)',r'\1',file_path)
    for s in wb.sheetnames:
        # どのファイルを指定しようと現在のパスに出力する
        F_out = open(pwd + '/' + file_name + '__' + s + '.txt','w',encoding = 'UTF8')
        sheet = wb[s]
        history = {}
        h_line = ''
        for r in range(1,sheet.max_row):
            line   = ''
            for c in range(1,sheet.max_column):
                value = '"' + fix_unSJIS(sheet.cell(r,c).value).replace('_x000D_','').replace('None','').replace('"','゛') + '"'
                
                if len(value) < 3:
                    # 値がない場合
                    if r == 1:
                        # 一行目で値がない場合は空白を返す
                        history[c] = ''
                    else:
                        # 値がなければ一行上の値を参照
                        value = history[c]
                else:
                    # 値があれば値を履歴に登録
                    history[c] = value
                    
                
                # 結合して一行を作成
                line += value + '\t'
                
            #print((line +"\t"+h_line).replace('\n',''))
            if line == h_line:
                # 履歴と同じ行になるのなら出力を回避
                #pass
                continue
            
            print(line)
            F_out.write(line + '\n')
            h_line = line
            

def fix_unSJIS(tar_str):
    return str(tar_str).encode('SJIS','ignore').decode('SJIS')

def fix_unUTF8(tar_str):
    return str(tar_str).encode('UTF8','ignore').decode('UTF8')

if __name__ == '__main__':
    usage = '%prog [options] <input_Excel_file1>\n'
    parser = OptionParser(usage=usage)
    (options, args) = parser.parse_args()

    if len(args) == 1:
        file_path = args[0]
        export_all(file_path)
    else:
        parser.error('Check the usage and arguments!')
